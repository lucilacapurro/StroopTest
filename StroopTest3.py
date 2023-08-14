import pygame
from pygame.locals import *
import time
import os
import pandas as pd
from openpyxl import load_workbook
import sys

########################################################################################################################

# Absolute path:
actual_dir= os.path.abspath(os.path.dirname(__file__))

# Alarm file path
alarm_file = 'alarm_beep.wav'
path_alarm_file = os.path.join(actual_dir, alarm_file) # alarm wav sound file relative path

# Baremo file path
baremo_file = 'Baremo.xlsx'
path_baremo_file = os.path.join(actual_dir, baremo_file) # baremo file relative path
# Baremo PC
PCsheet = "PC"
df_baremoPC = pd.read_excel(path_baremo_file, sheet_name=PCsheet)
# Baremo I
Isheet = "I" 
df_baremoI = pd.read_excel(path_baremo_file, sheet_name=Isheet) 

# PRE Registers file path:
PRE_excel_file = 'RegistrosPRE.xlsx'
path_PRE_excel_file = os.path.join(actual_dir, PRE_excel_file) # PRE registers excel file relative path

# POST Registers file path:
POST_excel_file = 'RegistrosPOST.xlsx'
path_POST_excel_file = os.path.join(actual_dir, POST_excel_file) # POST registers excel file relative path

########################################################################################################################

class stroopTest3:
    def __init__(self, screen, screen_width, screen_height, TEST, test_data):
        self.screen = screen
        self.screen_width = screen_width
        self.screen_height = screen_height
        self.clock = pygame.time.Clock()

        self.button_size_x = 130
        self.button_size_y = 30
        self.button_padding_x = 10
        self.button_padding_y = 30
        self.buttons = []

        self.current_test = TEST # PRE or POST 
        self.participant_code = test_data[0]
        self.participant_age = test_data[1]

        self.initial_time = 5  # countdown begins at 45 secs
        self.elapsed_time = 0
        self.clock_running = False
        self.running = False
        self.test_ended = False
        self.show_message = False
        self.score = None
        self.selected_button = None
        self.selected_x2 = False
        self.PCscore = None
        if int(self.participant_age) < 65: 
            self.age_correction = 5 # age correction for PC score for < 65 adults 
        else:
            self.age_correction = 15 # age correction for PC score for >= 65 adults 
        self.TPCscore = None
        self.test_data = test_data
        self.selected_save = False

        pygame.mixer.init()
        self.alarm_sound = pygame.mixer.Sound(path_alarm_file)  # alarm sound

    def create_buttons(self):
        column_texts = [
            ["VERDE", "ROJO", "AZUL", "VERDE", "ROJO", "AZUL", "ROJO", "VERDE", "AZUL", "ROJO"], 
            ["VERDE", "AZUL", "ROJO", "VERDE", "ROJO", "VERDE", "AZUL", "ROJO", "VERDE", "AZUL"],  
            ["AZUL", "VERDE", "AZUL", "ROJO", "VERDE", "AZUL", "VERDE", "ROJO", "VERDE", "ROJO"], 
            ["AZUL", "ROJO", "VERDE", "AZUL", "VERDE", "AZUL", "ROJO", "VERDE", "AZUL", "ROJO"], 
            ["VERDE", "ROJO", "VERDE", "AZUL", "ROJO", "VERDE", "ROJO", "AZUL", "ROJO", "AZUL"],
            ["VERDE", "AZUL", "ROJO", "VERDE", "AZUL", "VERDE", "ROJO", "AZUL", "ROJO", "VERDE"],
            ["ROJO", "AZUL", "ROJO", "VERDE", "AZUL", "VERDE", "AZUL", "ROJO", "AZUL", "ROJO"],
            ["VERDE", "ROJO", "VERDE", "AZUL", "VERDE", "ROJO", "AZUL", "ROJO", "VERDE", "AZUL"],
            ["VERDE", "ROJO", "VERDE", "ROJO", "AZUL", "ROJO", "AZUL", "VERDE", "ROJO", "VERDE"],
            ["AZUL", "VERDE", "AZUL", "ROJO", "AZUL", "ROJO", "VERDE", "AZUL", "VERDE", "ROJO"],
        ]

        for row in range(10):
            for col in range(10):
                x = col * (self.button_size_x + self.button_padding_x) + 20
                y = row * (self.button_size_y + self.button_padding_y) + 170
                word_button = pygame.Rect(x, y, self.button_size_x, self.button_size_y)
                button = pygame.draw.rect(self.screen, (255, 255, 255), word_button)
                text = column_texts[col][row]
                score = str(col * 10 + row + 1)
                self.buttons.append((button, text, row, col, score))

    def draw_ui(self):
        self.screen.fill((255, 255, 255))

        # Task name 
        font = pygame.font.SysFont(None, 45)
        task_text = "Test de Stroop 3"
        text_surface = font.render(task_text, True, (0, 0, 0))
        self.screen.blit(text_surface, (20, 20))

        # Draw clock
        font = pygame.font.SysFont(None, 36)
        if self.initial_time > self.elapsed_time > 0: 
            time_color = (255, 255, 255) # white (invisible) during countdown
        elif self.elapsed_time >= self.initial_time: 
            time_color = (255, 0, 0) # red when count down reaches zero
        else:
            time_color = (0, 0, 0)  # black before starting the countdown at 45 secs 
        time_text = "Tiempo: {:.0f}".format(max(self.initial_time - self.elapsed_time, 0))  # counts down
        text_surface = font.render(time_text, True, time_color)  # sets defined color
        self.screen.blit(text_surface, (self.screen_width/2-100, 50))

        # Draw clock start/reset button
        button_start_reset = pygame.Rect(self.screen_width/2+50, 50, self.button_size_x, self.button_size_y)
        pygame.draw.rect(self.screen, (0, 0, 0), button_start_reset)
        font = pygame.font.SysFont(None, 24)
        button_text = "Empezar" if not self.clock_running else "Reiniciar"
        text_surface = font.render(button_text, True, (255, 255, 255))
        text_rect = text_surface.get_rect(center=button_start_reset.center)
        self.screen.blit(text_surface, text_rect)

        # Draw x2 button
        button_x2 = pygame.Rect(self.screen_width/2+350, 50, 50, 30)
        if self.selected_x2: 
            pygame.draw.rect(self.screen, (255, 0, 0), button_x2)
        else: 
            pygame.draw.rect(self.screen, (0, 0, 0), button_x2)
        font = pygame.font.SysFont(None, 24)
        button_text = "x2"
        text_surface = font.render(button_text, True, (255, 255, 255))
        text_rect = text_surface.get_rect(center=button_x2.center)
        self.screen.blit(text_surface, text_rect)

        # Draw save button
        button_save = pygame.Rect(self.screen_width/2+450, 50, 100, 30)
        pygame.draw.rect(self.screen, (0, 0, 0), button_save)
        font = pygame.font.SysFont(None, 24)
        button_text = "Guardar"
        text_surface = font.render(button_text, True, (255, 255, 255))
        text_rect = text_surface.get_rect(center=button_save.center)
        self.screen.blit(text_surface, text_rect)

        # Draw buttons
        red = (255, 0, 0)
        green = (0, 205, 50)
        blue = (0, 0, 255)
        column_colors = [
            [red, blue, green, red, blue, green, blue, red, green, blue], 
            [red, green, blue, red, blue, red, green, blue, red, green],  
            [green, red, green, blue, red, green, red, blue, red, blue], 
            [green, blue, red, green, red, green, blue, red, green, blue],
            [red, blue, red, green, blue, red, blue, green, blue, green],
            [red, green, blue, red, green, red, blue, green, blue, red],
            [blue, green, blue, red, green, red, green, blue, green, blue],
            [red, blue, red, green, red, blue, green, blue, red, green],
            [red, blue, red, blue, green, blue, green, red, blue, red],
            [green, red, green, blue, green, blue, red, green, red, blue],
        ]

        for button, text, row, col, score in self.buttons:
            pygame.draw.rect(self.screen, (255, 255, 255), button)
            font = pygame.font.SysFont(None, 24)
            if button == self.selected_button and self.test_ended: 
                text_surface = font.render(text, True, (0, 0, 0))
            else:
                color = column_colors[col][row]
                text_surface = font.render(text, True, color)
            text_rect = text_surface.get_rect(center=button.center)
            self.screen.blit(text_surface, text_rect)

        # Show End of the Test message
        if self.show_message:
            font = pygame.font.SysFont(None, 30)
            message_text = "Fin del Test 1. Presione la palabra hasta dónde llegó el participante. Si completó una vuelta además presione x2. Luego presione Guardar."
            text_surface = font.render(message_text, True, (255, 0, 0))
            text_rect = text_surface.get_rect(center=(self.screen_width // 2, 120))
            self.screen.blit(text_surface, text_rect)

        pygame.display.flip()


    # logout message end of test
    def end_screen(self):
        self.screen.fill((255, 255, 255))
        self.text_module.show_message("Ingrese la edad del sujeto", big_font=False, no_bkg=True)


    def run(self):
        self.create_buttons()
        self.running = True

        mx, my = pygame.mouse.get_pos()

        while self.running:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    self.running = False

                if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:  # click on buttons
                    
                    button_start_reset = pygame.Rect(self.screen_width/2+50, 50, 130, 30)
                    if button_start_reset.collidepoint(event.pos): # click on start_button
                        self.clock_running = not self.clock_running # toggles the clock condition starts/reset
                        self.start_time = time.time()
                        if not self.clock_running: 
                            self.elapsed_time = 0  # resets the count down
                            self.start_time = time.time()
                    
                    for button, text, row, col, score in self.buttons:
                        if button.collidepoint(event.pos):  # click on buttons
                            self.selected_button = button 
                            self.score = score
                    
                    button_x2 = pygame.Rect(self.screen_width/2+350, 50, 50, 30)
                    if button_x2.collidepoint(event.pos): # click on x2_button
                        if self.test_ended: 
                            self.selected_x2 = not self.selected_x2
                    
                    button_save = pygame.Rect(self.screen_width/2+450, 50, 100, 30)
                    if button_save.collidepoint(event.pos): # click on save_button
                        if self.test_ended:
                            if self.selected_button != None: # allows saving if a word button is selected
                                self.selected_save = not self.selected_save
                                if self.selected_x2: # considers the x2 button
                                    self.PCscore = int(self.score) + 100 # calculates the P score
                                else:
                                    self.PCscore = int(self.score)

                                if self.selected_save:
                                    # raw PC score 
                                    print("Puntaje PC:", self.PCscore)
                                    self.test_data.append(self.PCscore)

                                    # age correction
                                    print("PC corregido por edad:", self.PCscore + self.age_correction)
                                    self.test_data.append(self.PCscore + self.age_correction)

                                    # TP score 
                                    try: 
                                        self.TPCscore = df_baremoPC.loc[df_baremoPC['PC'] == int(self.PCscore + self.age_correction), 'TPC'].values[0]
                                        print("Puntaje TPC:", self.TPCscore)
                                    except: 
                                        self.TPCscore = "-"
                                        print("El puntaje TPC no pudo ser calculado porque no se encontró el valor PC en el baremo.")
                                    self.test_data.append(self.TPCscore)

                                    # PC'
                                    try:
                                        PCestimated = round((self.test_data[3] * self.test_data[6]) / (self.test_data[3] + self.test_data[6]), 1) # PC' = (Pcorregido * Ccorregido) / (Pcorregido + Ccorregido)
                                        print("PC':", PCestimated)
                                    except: 
                                        PCestimated = "-"
                                        print("PC' no pudo ser calculado porque alguno de los puntajes P o C no pudo ser calculado correctamente:", PCestimated)
                                    self.test_data.append(PCestimated)

                                    # I score
                                    try: 
                                        Iscore = round(self.test_data[9] - PCestimated, 1) # I = real - estimated = PCcorregido - PC' 
                                        print("Puntaje I:", Iscore)
                                    except: 
                                        Iscore = "-"
                                        print("El puntaje I no pudo ser calculado porque no se pudo calcular el valor de TPC.")
                                    self.test_data.append(Iscore)

                                    # TI score
                                    try: 
                                        TIscore = df_baremoI.loc[df_baremoI['I'] == int(Iscore), 'TI'].values[0]
                                        print("Puntaje TI:", TIscore)
                                    except: 
                                        TIscore = "-"
                                        print("El puntaje TI no pudo ser calculado porque no se encontró el valor I en el baremo.")
                                    self.test_data.append(TIscore)

                                    # Saves the data 

                                    # PRE TEST CASE:
                                    if self.current_test == "PRE":
                                        '''
                                        # participant's data 
                                        "codigo": self.test_data[0], 
                                        "edad": self.test_data[1], 
                                        "P": self.test_data[2],
                                        "Pcorregido": self.test_data[3],
                                        "TP": self.test_data[4],
                                        "C": self.test_data[5],
                                        "Ccorregido": self.test_data[6],
                                        "TC": self.test_data[7],
                                        "PC": self.test_data[8],
                                        "PCcorregido": self.test_data[9],
                                        "TPC": self.test_data[10],
                                        "PC'": self.test_data[11],
                                        "I": self.test_data[12],
                                        "TI": self.test_data[13] 
                                        '''
                                        # saves the data in the PRE register 
                                        PRE_registers = load_workbook(path_PRE_excel_file)
                                        PRE_registers_sheet = PRE_registers["Hoja1"]

                                        # finds the first empty row
                                        first_empty_row = None
                                        for row in PRE_registers_sheet.iter_rows(min_row=1, max_row=PRE_registers_sheet.max_row):
                                            if all(cell.value is None for cell in row):
                                                first_empty_row = row[0].row
                                                break
                                        if first_empty_row is None:
                                            first_empty_row = PRE_registers_sheet.max_row + 1
                                        for cell, value in zip(PRE_registers_sheet[first_empty_row], self.test_data):
                                            cell.value = value

                                        PRE_registers.save(path_PRE_excel_file)

                                    # POST TEST CASE: 
                                    elif self.current_test == "POST":
                                        # gets the participant's PRE data from the PRE register 
                                        df_PRE_registers = pd.read_excel(path_PRE_excel_file)
                                        participant_data = list(df_PRE_registers.loc[df_PRE_registers['codigo'] == str(self.test_data[0])].iloc[0])

                                        # appends the participant's POST data
                                        participant_data.extend(self.test_data[2:])

                                        # compares PRE and POST values 
                                        # P
                                        try:
                                            Pdif = participant_data[14] - participant_data[2] # Pdif = Ppost - Ppre
                                            participant_data.append(Pdif)
                                            if Pdif < 10:
                                                Psig = 0 
                                            else:
                                                Psig = 1 # significant difference
                                        except: 
                                            Pdif = "-"
                                            Psig = "-"
                                        participant_data.append(Psig)

                                        # C
                                        try: 
                                            Cdif = participant_data[17] - participant_data[5] # Cdif = Cpost - Cpre
                                            participant_data.append(Cdif)
                                            if Cdif < 10:
                                                Csig = 0
                                            else:
                                                Csig = 1 # significant difference
                                        except: 
                                            Cdif = "-"
                                            Csig = "-"                           
                                        participant_data.append(Csig)                 

                                        # PC
                                        try:
                                            PCdif = participant_data[20] - participant_data[8] # PCdif = PCpost - PCpre
                                            participant_data.append(PCdif)
                                            if PCdif < 10:
                                                PCsig = 0
                                            else:
                                                PCsig = 1 # significant difference
                                        except:
                                            PCdif = "-"
                                            PCsig = "-"
                                        participant_data.append(PCsig)

                                        '''
                                        # participant's data 
                                        "SUJETO - codigo": participant_data[0], 
                                        "SUJETO - edad": participant_data[1], 
                                        "PRE - P": participant_data[2],
                                        "PRE - Pcorregido": participant_data[3],
                                        "PRE - TP": participant_data[4],
                                        "PRE - C": participant_data[5],
                                        "PRE - Ccorregido": participant_data[6],
                                        "PRE - TC": participant_data[7],
                                        "PRE - PC": participant_data[8],
                                        "PRE - PCcorregido": participant_data[9],
                                        "PRE - TPC": participant_data[10],
                                        "PRE - PC'": participant_data[11],
                                        "PRE - I": participant_data[12],
                                        "PRE - TI": participant_data[13],
                                        "POST - P": participant_data[14],
                                        "POST - Pcorregido": participant_data[15],
                                        "POST - TP": participant_data[16],
                                        "POST - C": participant_data[17],
                                        "POST - Ccorregido": participant_data[18],
                                        "POST - TC": participant_data[19],
                                        "POST - PC": participant_data[20],
                                        "POST - PCcorregido": participant_data[21],
                                        "POST - TPC": participant_data[22],
                                        "POST - PC'": participant_data[23],
                                        "POST - I": participant_data[24],
                                        "POST - TI": participant_data[25],
                                        "COMPARACION - DifP": participant_data[26],
                                        "COMPARACION - SigP": participant_data[27],
                                        "COMPARACION - DifC": participant_data[28],
                                        "COMPARACION - SigC": participant_data[29],
                                        "COMPARACION - DifPC": participant_data[30],
                                        "COMPARACION - SigPC": participant_data[31]
                                        '''
                                        
                                        # saves the data in the POST register 
                                        POST_registers = load_workbook(path_POST_excel_file)
                                        POST_registers_sheet = POST_registers["Hoja1"]

                                        # finds the first empty row
                                        first_empty_row = None
                                        for row in POST_registers_sheet.iter_rows(min_row=1, max_row=POST_registers_sheet.max_row):
                                            if all(cell.value is None for cell in row):
                                                first_empty_row = row[0].row
                                                break
                                        if first_empty_row is None:
                                            first_empty_row = POST_registers_sheet.max_row + 1
                                        for cell, value in zip(POST_registers_sheet[first_empty_row], participant_data):
                                            cell.value = value

                                        POST_registers.save(path_POST_excel_file)

                                    self.show_message = False
                                    self.screen.fill((255, 255, 255))  # Fill the screen with white color to clear the content
                                    font = pygame.font.SysFont(None, 30)
                                    message_text = "Fin del test. Los datos fueron guardados. A continuación se cerrará el programa" # end of test message
                                    text_surface = font.render(message_text, True, (0, 0, 0))  # Green color for success
                                    text_rect = text_surface.get_rect(center=(self.screen_width // 2, 150))
                                    self.screen.blit(text_surface, text_rect)
                                    pygame.display.flip()
                                    pygame.time.wait(4000)  # Waits for 4 seconds to show the message

                                    sys.exit() # exits the program

            if self.clock_running:
                self.elapsed_time = time.time() - self.start_time
                if self.elapsed_time >= self.initial_time: # countdown reaches 0
                    self.clock_running = False
                    self.show_message = True # shows end of the test message
                    self.alarm_sound.play() # plays alarm beep
                    self.test_ended = True

            self.draw_ui()
            self.clock.tick(60)
        
        pygame.display.flip()


                